from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
import io
import os
import re
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB

# ── CONSTANTS ─────────────────────────────────────────────────────────────────

PRIMARY_ACCOUNTS = {
    'Slash', 'Divvy Credit', 'Divvy Prefund',
    'Wex Credit', 'Wex Prefund', 'Global Rewards', 'Taekus',
}
EXCLUDE_ACCOUNTS = {'Clearing Account', 'Accounts Payable'}

DARK_BLUE = '1F3864'; MED_BLUE = '2E5EAA'; TEAL = '1F6B75'; GREEN = '1F6B3B'
ALT_ROW = 'EEF4FB'; ALT_OTHER = 'F0F7F4'; ALT_TEAM = 'F0F7F0'; WHITE = 'FFFFFF'


# ── VALIDATION ────────────────────────────────────────────────────────────────

def validate_gl(df):
    ncols = df.shape[1]
    if ncols < 9:
        return False, f"This looks like a previously generated report ({ncols} columns). Please upload the raw QuickBooks GL export."
    for col in [8, 9]:
        if col < ncols and pd.to_numeric(df[col], errors='coerce').notna().sum() > 5:
            return True, None
    return False, "This doesn't look like a General Ledger export. Please upload the raw QuickBooks GL export."


# ── FORMAT DETECTION ─────────────────────────────────────────────────────────

def detect_format(df):
    if 9 not in df.columns or 8 not in df.columns:
        return 'old'
    mask = df[9].notna() & df[0].isna() & df[1].notna() & \
           ~df[1].astype(str).str.contains('Beginning Balance|Transaction date', case=False, na=False)
    sample = df[mask][8].dropna()
    if len(sample) > 0 and sample.nunique() <= 20 and (len(sample) / max(sample.nunique(), 1)) > 5:
        try:
            pd.to_numeric(sample, errors='raise')
        except (ValueError, TypeError):
            return 'consolidated'
    return 'new' if str(df.iloc[1, 0]).strip() == 'Transaction Report' else 'old'


# ── DATE HELPERS ──────────────────────────────────────────────────────────────

def ordinal(n):
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th') if n not in (11, 12, 13) else 'th'

def next_month_date_label(date_range_str):
    actual = date_range_str.strip()
    m = re.search(r'(\w+)\s+(\d+)\w*[-–](\d+)\w*,?\s*(\d{4})', actual)
    if m:
        month_str, start_day, end_day, year_str = m.group(1), int(m.group(2)), int(m.group(3)), m.group(4)
        nxt = pd.to_datetime(f'{month_str} 1 {year_str}') + pd.offsets.MonthBegin(1)
        return f'{nxt.strftime("%B")} {start_day}{ordinal(start_day)}-{end_day}{ordinal(end_day)}', actual
    mm = re.search(r'(\w+)\s+(\d+)', actual)
    ym = re.search(r'(\d{4})', actual)
    month_str = mm.group(1)
    start_day = int(mm.group(2)) if ym and mm.group(2) != ym.group(1) else 1
    year_str  = ym.group(1) if ym else None
    nxt = pd.to_datetime(f'{month_str} 1 {year_str}') + pd.offsets.MonthBegin(1)
    return f'{nxt.strftime("%B")} {start_day}{ordinal(start_day)}', actual


# ── METADATA ─────────────────────────────────────────────────────────────────

def smart_title(s):
    ALLCAPS = {'llc','tl','ys','yss','gl','mls','nba','nhl','mlb','nfl',
               'tc','dep','lp','inc','ltd','kg','sp','yskg','ysp','ysm','gr'}
    result = []
    for w in s.split():
        if '&' in w:    result.append(w.upper() if len(w) <= 3 else w.title())
        elif w.lower() in ALLCAPS: result.append(w.upper())
        else:           result.append(w.capitalize())
    return ' '.join(result)

def build_meta(df, fmt):
    if fmt == 'consolidated':
        name           = 'All Companies'
        date_range_str = str(df.iloc[1].dropna().tolist()[0]).strip()
    elif fmt == 'old':
        name           = str(df.iloc[1].dropna().tolist()[0]).strip()
        date_range_str = str(df.iloc[2].dropna().tolist()[0]).strip()
    else:
        name           = str(df.iloc[0].dropna().tolist()[0]).strip()
        date_range_str = str(df.iloc[2].dropna().tolist()[0]).strip()
    next_label, actual_label = next_month_date_label(date_range_str)
    company_title = smart_title(name)
    report_title  = f'{company_title} - Installment Projections - {next_label}'
    return company_title, actual_label, next_label, report_title, f'{report_title}.xlsx'


# ── ACCOUNT RELABELING ────────────────────────────────────────────────────────

def relabel(account):
    a = str(account).strip(); al = a.lower()
    if 'slash plat' in al or al.startswith('sp') or ' sp' in al: return 'Slash'
    if 'divvy cr' in al or al in ('divvy (credit)', 'divvy credit'):  return 'Divvy Credit'
    if 'divvy pf' in al or al == 'divvy (prefund)':                   return 'Divvy Prefund'
    if 'wex (prefund)' in al or 'wex prefund' in al:                  return 'Wex Prefund'
    if 'wex cr' in al or 'wex (credit' in al:                         return 'Wex Credit'
    if 'global reward' in al or al.startswith('gr ') or al == 'gr':   return 'Global Rewards'
    return a


# ── DATA LOADING ──────────────────────────────────────────────────────────────

def load_transactions(df, fmt):
    if fmt == 'old':
        amt_col=8; acct_col=7; type_col=2; name_col=4; desc_col=5; date_col=1; comp_col=None
        mask = df[amt_col].notna() & (df[amt_col] != 'Amount') & df[0].isna()
    elif fmt == 'new':
        amt_col=9; acct_col=8; type_col=2; name_col=5; desc_col=6; date_col=1; comp_col=None
        mask = (df[amt_col].notna() & df[0].isna() & df[date_col].notna() &
                ~df[date_col].astype(str).str.contains('Beginning Balance|Date', case=False, na=False))
    else:
        amt_col=9; acct_col=7; type_col=2; name_col=4; desc_col=5; date_col=1; comp_col=8
        mask = (df[amt_col].notna() & df[0].isna() & df[date_col].notna() &
                ~df[date_col].astype(str).str.contains('Beginning Balance|Transaction date', case=False, na=False))
    tx = df[mask].copy()
    tx[amt_col] = pd.to_numeric(tx[amt_col], errors='coerce')
    tx = tx.dropna(subset=[amt_col])
    tx['labeled'] = tx[acct_col].apply(relabel)
    tx = tx[tx[type_col].notna() & (tx[type_col].astype(str).str.strip() != '')]
    tx = tx[~tx['labeled'].isin(EXCLUDE_ACCOUNTS)]
    tx = tx[~tx[acct_col].isin(EXCLUDE_ACCOUNTS)]
    cols  = [date_col, type_col, name_col, desc_col, 'labeled', amt_col]
    names = ['Date', 'Type', 'Name', 'Description', 'Account', 'Amount']
    if comp_col is not None:
        cols.append(comp_col); names.append('Company')
    result = tx[cols].copy(); result.columns = names
    result['Date'] = pd.to_datetime(result['Date'], errors='coerce')
    return result[result['Amount'] != 0]


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────

def hfont(sz=11): return Font(name='Arial', bold=True, color=WHITE, size=sz)
def cfont(sz=10): return Font(name='Arial', size=sz)
def tborder():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)
def bold_left_border():
    m = Side(style='medium', color='BFBFBF'); t = Side(style='thin', color='BFBFBF')
    return Border(left=m, right=t, top=t, bottom=t)


# ── LAYOUT HELPERS ────────────────────────────────────────────────────────────

def write_title(ws, report_title, actual_date_range, ncols):
    span = get_column_letter(ncols)
    ws.merge_cells(f'A1:{span}1'); c = ws['A1']; c.value = report_title
    c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f'A2:{span}2'); c = ws['A2']; c.value = actual_date_range
    c.font = Font(name='Arial', italic=True, color=WHITE, size=10)
    c.fill = PatternFill('solid', fgColor=MED_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

def write_sec_hdr(ws, row, label, color, c1, c2):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=label)
    c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = tborder(); ws.row_dimensions[row].height = 20

def write_col_hdrs(ws, row, hdrs, bg, c1=1):
    for i, h in enumerate(hdrs):
        c = ws.cell(row=row, column=c1+i, value=h)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = tborder()
    ws.row_dimensions[row].height = 20


# ── ACCOUNT SUMMARY (cols 1-4) ────────────────────────────────────────────────

def write_account_summary(ws, tx, start_row):
    acct = tx.groupby('Account')['Amount'].agg(['sum', 'count']).reset_index()
    acct.columns = ['Account', 'Total Spent', 'Transactions']
    acct = acct.sort_values('Total Spent', ascending=False).reset_index(drop=True)
    prim = acct[acct['Account'].isin(PRIMARY_ACCOUNTS)].sort_values('Total Spent', ascending=False).reset_index(drop=True)
    oth  = acct[~acct['Account'].isin(PRIMARY_ACCOUNTS)].sort_values('Total Spent', ascending=False).reset_index(drop=True)
    grand_total = float(tx['Amount'].sum())

    r = start_row
    write_sec_hdr(ws, r, '▸  Primary Accounts', MED_BLUE, 1, 4); r += 1
    write_col_hdrs(ws, r, ['Account', 'Total Spent ($)', '# Trans', '% Of Total'], '4472C4', 1); r += 1

    for i, row in prim.iterrows():
        fill = ALT_ROW if i % 2 == 0 else WHITE
        pct  = row['Total Spent'] / grand_total if grand_total else 0
        for col, (val, fmt, aln) in enumerate(zip(
            [row['Account'], row['Total Spent'], int(row['Transactions']), pct],
            [None, '$#,##0.00', '#,##0', '0.0%'],
            ['left', 'center', 'center', 'center']
        ), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = cfont(); c.fill = PatternFill('solid', fgColor=fill)
            c.border = tborder(); c.alignment = Alignment(horizontal=aln, vertical='center')
            if fmt: c.number_format = fmt
        ws.row_dimensions[r].height = 18; r += 1

    p_total = float(prim['Total Spent'].sum())
    for col, (val, fmt, aln) in enumerate(zip(
        ['Subtotal — Primary', p_total, int(prim['Transactions'].sum()), p_total / grand_total if grand_total else 0],
        [None, '$#,##0.00', '#,##0', '0.0%'], ['left', 'center', 'center', 'center']
    ), 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', fgColor=MED_BLUE); c.border = tborder()
        c.alignment = Alignment(horizontal=aln, vertical='center')
        if fmt: c.number_format = fmt
    ws.row_dimensions[r].height = 20; r += 1
    ws.row_dimensions[r].height = 8; r += 1

    write_sec_hdr(ws, r, '▸  Other Accounts', TEAL, 1, 4); r += 1
    write_col_hdrs(ws, r, ['Account', 'Total Spent ($)', '# Trans', '% Of Total'], '2E8B8F', 1); r += 1

    for i, row in oth.iterrows():
        fill = ALT_OTHER if i % 2 == 0 else WHITE
        pct  = row['Total Spent'] / grand_total if grand_total else 0
        for col, (val, fmt, aln) in enumerate(zip(
            [row['Account'], row['Total Spent'], int(row['Transactions']), pct],
            [None, '$#,##0.00', '#,##0', '0.0%'], ['left', 'center', 'center', 'center']
        ), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = cfont(); c.fill = PatternFill('solid', fgColor=fill)
            c.border = tborder(); c.alignment = Alignment(horizontal=aln, vertical='center')
            if fmt: c.number_format = fmt
        ws.row_dimensions[r].height = 18; r += 1

    o_total = float(oth['Total Spent'].sum())
    for col, (val, fmt, aln) in enumerate(zip(
        ['Subtotal — Other', o_total, int(oth['Transactions'].sum()), o_total / grand_total if grand_total else 0],
        [None, '$#,##0.00', '#,##0', '0.0%'], ['left', 'center', 'center', 'center']
    ), 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', fgColor=TEAL); c.border = tborder()
        c.alignment = Alignment(horizontal=aln, vertical='center')
        if fmt: c.number_format = fmt
    ws.row_dimensions[r].height = 20; r += 1
    ws.row_dimensions[r].height = 8; r += 1

    for col, (val, fmt, aln) in enumerate(zip(
        ['Grand Total', grand_total, int(acct['Transactions'].sum()), 1.0],
        [None, '$#,##0.00', '#,##0', '0.0%'], ['left', 'center', 'center', 'center']
    ), 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=11)
        c.fill = PatternFill('solid', fgColor=DARK_BLUE); c.border = tborder()
        c.alignment = Alignment(horizontal=aln, vertical='center')
        if fmt: c.number_format = fmt
    ws.row_dimensions[r].height = 24; r += 1
    return r


# ── TEAM SUMMARY (cols 6-9) ───────────────────────────────────────────────────

def write_team_summary(ws, tx, start_row):
    team = tx.groupby('Name')['Amount'].agg(['sum', 'count']).reset_index()
    team.columns = ['Team', 'Total Spent', 'Transactions']
    team = team.sort_values('Total Spent', ascending=False).reset_index(drop=True)
    grand_total = float(tx['Amount'].sum())

    r = start_row
    write_sec_hdr(ws, r, f'▸  Spending By Team  ({len(team)} Teams)', GREEN, 6, 9); r += 1
    write_col_hdrs(ws, r, ['Team', 'Total Spent ($)', '# Trans', '% Of Total'], '2E8B3B', 6); r += 1

    for i, row in team.iterrows():
        fill = ALT_TEAM if i % 2 == 0 else WHITE
        pct  = row['Total Spent'] / grand_total if grand_total else 0
        for j, (val, fmt, aln) in enumerate(zip(
            [row['Team'], row['Total Spent'], int(row['Transactions']), pct],
            [None, '$#,##0.00', '#,##0', '0.0%'], ['left', 'center', 'center', 'center']
        )):
            c = ws.cell(row=r, column=6+j, value=val)
            c.font = cfont(); c.fill = PatternFill('solid', fgColor=fill)
            c.border = bold_left_border() if j == 0 else tborder()
            c.alignment = Alignment(horizontal=aln, vertical='center')
            if fmt: c.number_format = fmt
        ws.row_dimensions[r].height = 18; r += 1

    for j, (val, fmt, aln) in enumerate(zip(
        ['Grand Total', grand_total, int(team['Transactions'].sum()), 1.0],
        [None, '$#,##0.00', '#,##0', '0.0%'], ['left', 'center', 'center', 'center']
    )):
        c = ws.cell(row=r, column=6+j, value=val)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=11)
        c.fill = PatternFill('solid', fgColor=DARK_BLUE); c.border = tborder()
        c.alignment = Alignment(horizontal=aln, vertical='center')
        if fmt: c.number_format = fmt
    ws.row_dimensions[r].height = 24; r += 1
    return r


# ── COMPANY TAB ───────────────────────────────────────────────────────────────

def build_company_tab(wb, tx, report_title, actual_date_range, tab_name, is_first=False):
    ws = wb.active if is_first else wb.create_sheet(tab_name[:31])
    if is_first: ws.title = tab_name[:31]
    subtitle = f'{actual_date_range}  |  {len(tx):,} Transactions  |  ${tx["Amount"].sum():,.2f} Total'
    write_title(ws, report_title, subtitle, 9)
    acct_end = write_account_summary(ws, tx, 4)
    write_team_summary(ws, tx, 4)
    widths = {1:24, 2:16, 3:10, 4:12, 5:3, 6:30, 7:16, 8:10, 9:12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in range(4, acct_end + 80):
        c = ws.cell(row=row, column=5, value='')
        c.fill = PatternFill('solid', fgColor='F0F0F0')
    ws.freeze_panes = 'A4'


# ── TRANSACTIONS TAB ──────────────────────────────────────────────────────────

def build_transactions_tab(wb, tx, report_title, actual_date_range):
    tdf = tx.sort_values('Amount', ascending=False).reset_index(drop=True)
    has_company = 'Company' in tdf.columns
    ncols = 7 if has_company else 6
    ws = wb.create_sheet('Transactions')
    write_title(ws, report_title, f'{actual_date_range}  |  {len(tdf):,} Transactions', ncols)
    headers = ['Date', 'Type', 'Name', 'Description', 'Account', 'Amount ($)']
    if has_company: headers.append('Company')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hfont(); c.fill = PatternFill('solid', fgColor=MED_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = tborder()
    ws.row_dimensions[4].height = 22
    for i, row in tdf.iterrows():
        r = i + 5; fill = ALT_ROW if i % 2 == 0 else WHITE
        vals = [row['Date'].strftime('%m/%d/%Y') if pd.notna(row['Date']) else '',
                row['Type'], row['Name'], row['Description'], row['Account'], row['Amount']]
        if has_company: vals.append(row.get('Company', ''))
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = cfont(sz=9); c.fill = PatternFill('solid', fgColor=fill)
            c.border = tborder(); c.alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=r, column=6).number_format = '$#,##0.00'
        ws.row_dimensions[r].height = 16
    widths = [13, 16, 22, 52, 18, 14]
    if has_company: widths.append(22)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'


# ── MAIN REPORT PIPELINE ──────────────────────────────────────────────────────

def generate_report(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)
    valid, err = validate_gl(df)
    if not valid:
        raise ValueError(err)
    fmt = detect_format(df)
    company_title, actual_label, next_label, report_title, filename = build_meta(df, fmt)
    tx = load_transactions(df, fmt)

    wb = Workbook()
    if fmt == 'consolidated':
        companies = sorted(tx['Company'].dropna().unique(),
                           key=lambda x: tx[tx['Company']==x]['Amount'].sum(), reverse=True)
        all_title = f'All Companies - Installment Projections - {next_label}'
        build_company_tab(wb, tx, all_title, actual_label, 'All Companies', is_first=True)
        for company in companies:
            co_tx    = tx[tx['Company'] == company].copy()
            co_title = f'{smart_title(company)} - Installment Projections - {next_label}'
            build_company_tab(wb, co_tx, co_title, actual_label, smart_title(company))
        build_transactions_tab(wb, tx, all_title, actual_label)
        n_companies = len(companies)
    else:
        build_company_tab(wb, tx, report_title, actual_label, 'Summary', is_first=True)
        build_transactions_tab(wb, tx, report_title, actual_label)
        n_companies = 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    stats = {
        'filename': filename,
        'transactions': len(tx),
        'accounts': tx['Account'].nunique(),
        'teams': tx['Name'].nunique(),
        'total': float(tx['Amount'].sum()),
        'date_range': actual_label,
        'companies': n_companies,
    }
    return filename, buf.read(), stats


# ── FLASK ROUTES ──────────────────────────────────────────────────────────────

@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    try:
        files = request.files.getlist('files')
        if not files or all(f.filename == '' for f in files):
            return jsonify({'error': 'No files uploaded.'}), 400

        results = []
        for f in files:
            if not f.filename.endswith('.xlsx'):
                continue
            try:
                filename, xlsx_bytes, stats = generate_report(f.read())
                results.append({
                    'filename': filename,
                    'stats': stats,
                    'data': xlsx_bytes.hex(),  # send as hex to reconstruct on client
                })
            except Exception as e:
                results.append({'filename': f.filename, 'error': str(e)})

        if not results:
            return jsonify({'error': 'No valid Excel files found.'}), 400

        return jsonify({'results': results})

    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
